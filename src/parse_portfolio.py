# import pandas as pd
import os
import openpyxl as oxl
import xlrd
from datetime import datetime

import get_file_list as gfl
import sql_database as sq


def parse_trade_dict(return_trade_dict, ws, cur_row, xlsx_col, table_name):
    """Разбираем таблицы сделок. Их три:
    Сделки, совершенные с ЦБ на биржевых торговых площадках (Фондовый рынок) с расчетами в дату заключения;
    Сделки, совершенные с ЦБ на биржевых торговых площадках (Фондовый рынок) с расчетами Т+,
    незавершенные в отчетном периоде;
    Сделки, совершенные с ЦБ на биржевых торговых площадках (Фондовый рынок) с расчетами Т+,
    рассчитанные в отчетном периоде
    Возвращаем словарь return_trade_dict, который принимаемн а вход пустым
    """
    temp_row = 0
    not_cancel_col = 0
    if table_name == "незавершенные в отчетном периоде":
        not_cancel_col = 1

    while ws.cell(cur_row + temp_row + 2, xlsx_col + 1).value != "Итого оборот":
        # Дата и время совершения сделки. Плановая дата исполнения сделки в тот же день
        date_time_trade = datetime.strptime(ws.cell(cur_row + temp_row + 2, xlsx_col + 1).value,
                                            '%d.%m.%Y %H:%M:%S')
        return_trade_dict['date_time_trade'].append(date_time_trade.strftime('%d.%m.%Y %H:%M:%S'))
        # Плановая дата исполнения сделки
        if table_name == "незавершенные в отчетном периоде":
            date_time_exec = datetime.strptime(ws.cell(cur_row + temp_row + 2, xlsx_col + 2).value, '%d.%m.%Y')
            return_trade_dict['date_time_execution'].append(date_time_exec.strftime('%d.%m.%Y'))
        else:
            return_trade_dict['date_time_execution'].append(date_time_trade.strftime('%d.%m.%Y'))
        # Номер сделки в ТС
        number_trade = int(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 2).value)
        return_trade_dict['number_trade'].append(number_trade)
        # Наименование эмитента, вид, категория (тип), выпуск, транш ЦБ
        name_paper = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 4).value
        return_trade_dict['name_paper'].append(name_paper)
        # ISIN
        isin = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 7).value
        return_trade_dict['isin'].append(isin)
        # Номер гос. регистрации
        reg_num = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 8).value
        return_trade_dict['reg_num'].append(reg_num)
        # Вид сделки (покупка/продажа)
        type_trade = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 9).value
        return_trade_dict['type_trade'].append('buy' if type_trade == 'покупка' else 'sell')
        # Кол-во ЦБ, шт.
        volume = int(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 10).value)
        return_trade_dict['volume'].append(volume)
        # Цена (% для обл)
        transac_price = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 12).value)
        return_trade_dict['transac_price'].append(transac_price)
        # Сумма сделки без НКД
        transac_amount = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 13).value)
        return_trade_dict['transac_amount'].append(transac_amount)
        # НКД
        nkd = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 14).value)
        return_trade_dict['nkd'].append(nkd)
        # Комиссия торговой системы
        commission_TS = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 15).value)
        return_trade_dict['commission_TS'].append(commission_TS)
        # Клиринговая комиссия
        commission_klir = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 16).value)
        return_trade_dict['commission_klir'].append(commission_klir)
        # Комиссия за ИТС
        commission_its = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 17).value)
        return_trade_dict['commission_its'].append(commission_its)
        # Комиссия брокера
        commission_brok = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 19).value)
        return_trade_dict['commission_brok'].append(commission_brok)

        temp_row += 1

    return return_trade_dict


def parse_excel_report(file_name_parse):
    """Парсим файл с ежедневным отчетом и
     возвращаем словарь с активами assets_dict и словарь с портфелем portfolio_dict"""
    # Возвращаемые словари
    # Словарь активов на день
    return_assets_dict = {'time_report': None,
                          'incoming_amount': None,
                          'outgoing_amount': None,
                          'credit_customer': None,
                          'credit_corporate': None,
                          'assets_at_start': None,
                          'assets_at_end': None}
    # Таблица сделок
    return_trade_dict = {'date_time_trade': [],
                         'date_time_execution': [],
                         'number_trade': [],
                         'name_paper': [],
                         'isin': [],
                         'reg_num': [],
                         'type_trade': [],
                         'volume': [],
                         'transac_price': [],
                         'transac_amount': [],
                         'nkd': [],
                         'commission_TS': [],
                         'commission_klir': [],
                         'commission_its': [],
                         'commission_brok': []}
    # Состояние портфеля на день
    return_portfolio_dict = {}
    # Расширение файла
    file_ext = file_name_parse.split(".")[-1]

    if file_ext == "xls":
        workb = xlrd.open_workbook(file_name_parse)
        # Загружаем активный лист
        ws = workb.sheet_by_index(0)
        # Получаем максимальное количество заполненных строк
        ws_rows = ws.nrows
        xlsx_col = 0                # В openpyxl отсчет столбцов начинается с 1, в xlrd с 0
    elif file_ext == "xlsx":
        workb = oxl.load_workbook(filename=file_name_parse, read_only=True, data_only=True)
        # Загружаем активный лист
        ws = workb.active
        # Получаем максимальное количество заполненных строк
        ws_rows = ws.max_row
        xlsx_col = 1                # В openpyxl отсчет столбцов начинается с 1

    # Пробегаем файл по строчкам
    # Заполняем словарь активов на день
    cur_row = 1
    while cur_row < ws_rows:
        cell_name = "A" + str(cur_row)
        # Дата отчета
        if ws.cell(cur_row, xlsx_col).value == "ПериодДат":
            time_rep = datetime.strptime(ws.cell(cur_row, 3 + xlsx_col).value.split(" ")[-1], '%d.%m.%Y')  # In datetime
            return_assets_dict['time_report'] = time_rep.strftime('%d.%m.%Y')                        # In str
        # Входящая сумма средств на счете
        if ws.cell(cur_row, xlsx_col).value == "100":
            incoming_amount = ws.cell(cur_row, 5 + xlsx_col).value
            return_assets_dict['incoming_amount'] = incoming_amount
        # Начислено клиентом и в рамках корпоративных действий
        if ws.cell(cur_row, xlsx_col).value == "230":
            credit_customer = ws.cell(cur_row, 5 + xlsx_col).value
            credit_corporate = ws.cell(cur_row + 1, 5 + xlsx_col).value
            return_assets_dict['credit_customer'] = credit_customer
            return_assets_dict['credit_corporate'] = credit_corporate
        # Исходящая сумма средств на счете
        if ws.cell(cur_row, xlsx_col).value == "2250":
            outgoing_amount = ws.cell(cur_row, 5 + xlsx_col).value
            return_assets_dict['outgoing_amount'] = outgoing_amount
        # Сумма активов на начало и конец дня
        if ws.cell(cur_row, xlsx_col).value == "2600":
            assets_at_start = ws.cell(cur_row, 5 + xlsx_col).value
            assets_at_end = ws.cell(cur_row + 1, 5 + xlsx_col).value
            return_assets_dict['assets_at_start'] = assets_at_start
            return_assets_dict['assets_at_end'] = assets_at_end
            break                     # Выходим так как достигли конца таблицы со средствами

        cur_row += 1
    # В старых отчетах исходящий остаток не приводился, присваиваем как входящий остаток
    if return_assets_dict['incoming_amount'] and not return_assets_dict['outgoing_amount']:
        return_assets_dict['outgoing_amount'] = return_assets_dict['incoming_amount']

    # Разбираем таблицы сделок
    while cur_row < ws_rows:
        if str(ws.cell(cur_row, xlsx_col + 1).value).endswith('с расчетами в дату заключения'):
            return_trade_dict = parse_trade_dict(return_trade_dict,
                                                 ws,
                                                 cur_row,
                                                 xlsx_col,
                                                 table_name='с расчетами в дату заключения')

        if str(ws.cell(cur_row, xlsx_col + 1).value).endswith('незавершенные в отчетном периоде') \
                or str(ws.cell(cur_row, xlsx_col + 1).value).endswith('рассчитанные в отчетном периоде'):
            return_trade_dict = parse_trade_dict(return_trade_dict,
                                                 ws,
                                                 cur_row,
                                                 xlsx_col,
                                                 table_name='незавершенные в отчетном периоде')
        cur_row += 1

    print(return_trade_dict)

    return return_assets_dict, return_portfolio_dict


if __name__ == "__main__":
    # Получаем список всех файлов с отчетами
    file_list = gfl.get_file_list(os.getcwd() + "\\report\\")
    # file_list = gfl.get_file_list("d:\\olega\\Финансы\\Брокер\\Отчеты ПСБ\\")
    # print(file_list)

    db = sq.SQLiter("portfolio.db")
    db.create_table()

    for count, file_n in enumerate(file_list):
        assets_dict, portfolio_dict = parse_excel_report(file_n)
        # print(assets_dict)
        # db.insert_data(assets_dict)
        file_name = file_n.split("\\")[-1]
        print(f"Отчет № {str(count)} из {len(file_list)} - {file_name}")
    #
    db.close()
