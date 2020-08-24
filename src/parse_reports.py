import os
import openpyxl as oxl
import xlrd
from datetime import datetime

import get_file_list as gfl
import sql_database as sq

# TODO
# 1. Отрефакторить создание таблиц в БД и вставку данных
# 3. Составить портфель по сделкам


def list_of_dict():
    # Инициализация возвращаемых словарей
    # Словарь активов на день
    assets_dict_ = {'time_report': [],
                    'incoming_amount': [],
                    'outgoing_amount': [],
                    'credit_customer': [],
                    'credit_corporate': [],
                    'assets_at_start': [],
                    'assets_at_end': [],
                    'file_name': []}
    # Таблица сделок
    trade_dict_ = {'date_time_trade': [],
                   'date_time_execution': [],
                   'number_trade': [],
                   'name_paper': [],
                   'isin': [],
                   'reg_num': [],
                   'type_trade': [],
                   'volume': [],
                   'transact_price': [],
                   'transact_amount': [],
                   'nkd': [],
                   'commission_ts': [],
                   'commission_clear': [],
                   'commission_its': [],
                   'commission_brok': []}
    # Состояние портфеля на день
    portfolio_dict_ = {}

    return assets_dict_, trade_dict_, portfolio_dict_


def parse_assets_table(file_name_a, return_assets_dict, ws, ws_rows, xlsx_col):
    """Разбираем таблицу Сводная информация по счетам клиента в валюте счета
    Она одна и в находится в начале.
    Возвращаем словарь return_assets_dict, который принимаем на вход пустым
    """
    return_assets_dict['file_name'].append(file_name_a)
    cur_row = 1
    while cur_row < ws_rows:
        # Дата отчета
        if ws.cell(cur_row, xlsx_col).value == "ПериодДат" \
                or ws.cell(cur_row, 1 + xlsx_col).value == "ОТЧЕТ БРОКЕРА":
            time_rep = datetime.strptime(ws.cell(cur_row, 3 + xlsx_col).value.split(" ")[-1], '%d.%m.%Y')  # In datetime
            return_assets_dict['time_report'].append(time_rep.strftime('%d.%m.%Y'))  # In str
        # Входящая сумма средств на счете
        if ws.cell(cur_row, xlsx_col).value == "100" \
                or ws.cell(cur_row, 1 + xlsx_col).value == "ВХОДЯЩАЯ СУММА СРЕДСТВ НА СЧЕТЕ":
            incoming_amount = ws.cell(cur_row, 5 + xlsx_col).value
            return_assets_dict['incoming_amount'].append(incoming_amount)
        # Начислено клиентом и в рамках корпоративных действий
        if ws.cell(cur_row, xlsx_col).value == "200" \
                or ws.cell(cur_row, 1 + xlsx_col).value == "ЗАЧИСЛЕНО НА СЧЕТ":
            credit_customer = ws.cell(cur_row + 1, 5 + xlsx_col).value
            credit_corporate = ws.cell(cur_row + 2, 5 + xlsx_col).value
            return_assets_dict['credit_customer'].append(credit_customer)
            return_assets_dict['credit_corporate'].append(credit_corporate)
        # Исходящая сумма средств на счете
        if ws.cell(cur_row, xlsx_col).value == "2200" \
                or ws.cell(cur_row, 1 + xlsx_col).value == "ОСТАТОК СРЕДСТВ НА СЧЕТЕ":
            outgoing_amount = ws.cell(cur_row, 5 + xlsx_col).value
            return_assets_dict['outgoing_amount'].append(outgoing_amount)
        # Сумма активов на начало и конец дня
        if ws.cell(cur_row, xlsx_col).value == "2600" \
                or ws.cell(cur_row, 1 + xlsx_col).value == '"СУММА АКТИВОВ" на начало дня':
            assets_at_start = ws.cell(cur_row, 5 + xlsx_col).value
            assets_at_end = ws.cell(cur_row + 1, 5 + xlsx_col).value
            return_assets_dict['assets_at_start'].append(assets_at_start)
            return_assets_dict['assets_at_end'].append(assets_at_end)
            break  # Выходим так как достигли конца таблицы со средствами

        cur_row += 1

    return return_assets_dict


def parse_trade_tables(return_trade_dict, ws, cur_row, xlsx_col, table_name):
    """Разбираем таблицы сделок. Их три:
    Сделки, совершенные с ЦБ на биржевых торговых площадках (Фондовый рынок) с расчетами в дату заключения;
    Сделки, совершенные с ЦБ на биржевых торговых площадках (Фондовый рынок) с расчетами Т+,
    незавершенные в отчетном периоде;
    Сделки, совершенные с ЦБ на биржевых торговых площадках (Фондовый рынок) с расчетами Т+,
    рассчитанные в отчетном периоде
    Возвращаем словарь return_trade_dict, который принимаем на вход пустым
    """
    temp_row = 0
    not_cancel_col = 0
    if table_name == "незавершенные в отчетном периоде":
        not_cancel_col = 1

    while ws.cell(cur_row + temp_row + 2, xlsx_col + 1).value != "Итого оборот":
        # Дата и время совершения сделки. Плановая дата исполнения сделки в тот же день
        date_time_trade = datetime.strptime(ws.cell(cur_row + temp_row + 2, xlsx_col + 1).value,
                                            '%d.%m.%Y %H:%M:%S')
        return_trade_dict['date_time_trade'].append(date_time_trade.strftime('%d.%m.%Y %H:%M:%S'))
        # Плановая дата исполнения сделки
        if table_name == "незавершенные в отчетном периоде":
            date_time_exec = datetime.strptime(ws.cell(cur_row + temp_row + 2, xlsx_col + 2).value, '%d.%m.%Y')
            return_trade_dict['date_time_execution'].append(date_time_exec.strftime('%d.%m.%Y'))
        else:
            return_trade_dict['date_time_execution'].append(date_time_trade.strftime('%d.%m.%Y'))
        # Номер сделки в ТС
        number_trade = int(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 2).value)
        return_trade_dict['number_trade'].append(number_trade)
        # Наименование эмитента, вид, категория (тип), выпуск, транш ЦБ
        name_paper = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 4).value
        return_trade_dict['name_paper'].append(name_paper)
        # ISIN
        isin = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 7).value
        return_trade_dict['isin'].append(isin)
        # Номер гос. регистрации
        reg_num = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 8).value
        return_trade_dict['reg_num'].append(reg_num)
        # Вид сделки (покупка/продажа)
        type_trade = ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 9).value
        return_trade_dict['type_trade'].append('buy' if type_trade == 'покупка' else 'sell')
        # Кол-во ЦБ, шт.
        volume = int(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 10).value)
        return_trade_dict['volume'].append(volume)
        # Цена (% для обл)
        transact_price = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 12).value)
        return_trade_dict['transact_price'].append(transact_price)
        # Сумма сделки без НКД
        transact_amount = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 13).value)
        return_trade_dict['transact_amount'].append(transact_amount)
        # НКД
        nkd = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 14).value)
        return_trade_dict['nkd'].append(nkd)
        # Комиссия торговой системы
        commission_ts = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 15).value)
        return_trade_dict['commission_ts'].append(commission_ts)
        # Клиринговая комиссия
        commission_clear = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 16).value)
        return_trade_dict['commission_clear'].append(commission_clear)
        # Комиссия за ИТС
        commission_its = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 17).value)
        return_trade_dict['commission_its'].append(commission_its)
        # Комиссия брокера
        commission_brok = float(ws.cell(cur_row + temp_row + 2, xlsx_col + not_cancel_col + 19).value)
        return_trade_dict['commission_brok'].append(commission_brok)

        temp_row += 1

    return return_trade_dict


def parse_excel_report(file_name_parse):
    """Парсим файл с ежедневным отчетом и
     возвращаем словарь с активами return_assets_dict, словарь с портфелем return_portfolio_dict
     и словарь со сделками return_trade_dict"""

    # Расширение файла
    file_ext = file_name_parse.split(".")[-1]

    if file_ext == "xls":
        workb = xlrd.open_workbook(file_name_parse)
        # Загружаем активный лист
        ws = workb.sheet_by_index(0)
        # Получаем максимальное количество заполненных строк
        ws_rows = ws.nrows
        xlsx_col = 0                # В openpyxl отсчет столбцов начинается с 1, в xlrd с 0
    elif file_ext == "xlsx":
        workb = oxl.load_workbook(filename=file_name_parse, read_only=True, data_only=True)
        # Загружаем активный лист
        ws = workb.active
        # Получаем максимальное количество заполненных строк
        ws_rows = ws.max_row
        xlsx_col = 1                # В openpyxl отсчет столбцов начинается с 1

    return_assets_dict, return_trade_dict, return_portfolio_dict = list_of_dict()
    # Пробегаем файл по строчкам
    # Разбираем таблицу активов на день
    return_assets_dict = parse_assets_table(file_name_parse, return_assets_dict, ws, ws_rows, xlsx_col)

    # Разбираем таблицы сделок
    cur_row = 60
    while cur_row < ws_rows:
        if str(ws.cell(cur_row, xlsx_col + 1).value).endswith('с расчетами в дату заключения'):
            return_trade_dict = parse_trade_tables(return_trade_dict, ws,
                                                   cur_row, xlsx_col,
                                                   table_name='с расчетами в дату заключения')
            # Если прочитали в таблицу, то шагаем сразу на 6 строк (минимум в таблице)
            cur_row += 6
            continue

        if str(ws.cell(cur_row, xlsx_col + 1).value).endswith('незавершенные в отчетном периоде') \
                or str(ws.cell(cur_row, xlsx_col + 1).value).endswith('рассчитанные в отчетном периоде'):
            return_trade_dict = parse_trade_tables(return_trade_dict, ws,
                                                   cur_row, xlsx_col,
                                                   table_name='незавершенные в отчетном периоде')
            # Если прочитали в таблицу, то шагаем сразу на 6 строк (минимум в таблице)
            cur_row += 6
            continue

        cur_row += 1

    return return_assets_dict, return_trade_dict, return_portfolio_dict


if __name__ == "__main__":
    # Получаем список всех файлов с отчетами
    full_file_list = gfl.get_file_list(os.getcwd() + "\\report\\")
    # file_list = gfl.get_file_list("d:\\olega\\Финансы\\Брокер\\Отчеты ПСБ\\")
    # print(file_list)

    # Создаем базу данных если еще не создавалась
    db = sq.SQLiter("portfolio.db")
    # Создаем таблицы если еще не создавались
    assets_dict, trade_dict, _ = list_of_dict()
    db.create_table(assets_dict,
                    type_list=["CHAR(64)"] + ["FLOAT"] * 6 + ["CHAR"],
                    table_name='assets')
    db.create_table(trade_dict,
                    type_list=["CHAR(64)"] * 2 + ["BIGINT", "CHAR(256)"] + ["CHAR(64)"] * 3 + ["INT"] + ["FLOAT"] * 7,
                    table_name='trades')

    # Получаем список файлов уже хранящихся в базе данных
    exist_tuple = db.get_list_filename("assets")
    exist_files = [value for value, in exist_tuple]

    # Сравниваем два списка и оставляем только новые файлы которых нет в базе данных
    file_list = gfl.get_unique_file(full_file_list, exist_files)

    # Добавляем данные из файлов с отчетами в базу данных
    for count, file_n in enumerate(file_list):
        file_name = file_n.split("\\")[-1]
        print(f"Отчет № {str(count + 1)} из {len(file_list)} - {file_name}")

        assets_dict, trade_dict, portfolio_dict = parse_excel_report(file_n)
        # print(trade_dict)
        # print(assets_dict)

        db.insert_data(assets_dict, "assets")
        db.insert_data(trade_dict, "trades")

    db.close()
